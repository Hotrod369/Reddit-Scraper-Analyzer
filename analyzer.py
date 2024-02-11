'''
Analyzes data from scraper.py and finds potential bot accounts
'''
import json
import pandas as pd
import logging
import nltk
from nltk.sentiment import SentimentIntensityAnalyzer
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import datetime as dt  # Add this line to import the datetime module
from collections import defaultdict


logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()


# Load data from JSON files
with open("user_data.json", "r", encoding="utf-8") as user_file:
    user_data = json.load(user_file)

with open("submission_data.json", "r", encoding="utf-8") as submission_file:
    submission_data = json.load(submission_file)
    logger.info(f"loaded data from config files{len(user_data), len(submission_data)}")
    
# Initialize SentimentIntensityAnalyzer
sia = SentimentIntensityAnalyzer()

# Example function to calculate sentiment score of a text
def calculate_sentiment(text):
    score = sia.polarity_scores(text)
    return score['compound']  # Returns the compound score as sentiment


# Create a DataFrame from the user data
data = pd.DataFrame.from_dict(user_data, orient="index")
logger.info("Created a DataFrame from the user data.")

# Define thresholds for bot characteristics
KARMA_THRESHOLD = 5000
ACCOUNT_AGE_THRESHOLD = 2  # years

# Assuming config is undefined or None
config = None

# Create a default dictionary with an empty string as the default value
config = defaultdict(str) if config is None else config
logger.info(f"Created a default dictionary with an empty string as the default value{config}")

# Access the "notes" field with a default value
notes = config.get("notes", "")
if notes:
    notes = " ".join(notes)
    print("Note:", notes)

# Create a DataFrame to store potential bot accounts and their criteria
bot_df = pd.DataFrame(columns=["User", "Criteria", "Account Age", "Awardee Karma", "Link Karma",
                            "Comment Karma", "Total Karma", "User is Contributor",
                            "Has Verified Email", "Accepts Followers", "Average Sentiment Score"])
logger.info(f"Created a DataFrame{bot_df}")

# Calculate average sentiment score for comments in a submission
def calculate_average_sentiment(comments):
    if not comments:
        return 0  # Return neutral score if there are no comments
    scores = [sia.polarity_scores(comment['body'])['compound'] for comment in comments]
    average_score = sum(scores) / len(scores)
    logger.debug("Calculated average sentiment score for comments in a submission")
    return average_score

# Create a set to keep track of users already added to the DataFrame
added_users = set()

try:
    # Iterate through the submissions in submission_data and update user_data
    for submission_id, submission_info in submission_data.items():
        username = submission_info["User"]
        comments = submission_info.get("comments", [])
        average_sentiment_score = calculate_average_sentiment(comments)
        awardee_karma = submission_info.get("awardee_karma", 0)
        link_karma = submission_info.get("link_karma", 0)
        comment_karma = submission_info.get("comment_karma", 0)
        total_karma = submission_info.get("total_karma", 0)
        user_is_contributor = submission_info.get("user_is_contributor", False)
        has_verified_email = submission_info.get("has_verified_email", False)
        accept_followers = submission_info.get("accept_followers", False)
        logger.debug(f"Iterated through the submissions in submission_data{submission_id, submission_info}")

        
        # Fetch the user's data from user_data
        user_info = user_data.get(username, {})
        if not user_info:
            logger.info(f"Fetched the user's data from user_data")
            continue  # Skip if user data is not available
        
        # Calculate the account age using created_utc from user_data
        account_creation_time = dt.datetime.utcfromtimestamp(user_info.get("created_utc", 0))
        current_time = dt.datetime.utcnow()
        account_age = (current_time - account_creation_time).days / 365.25
        logger.info("Calculated the account users ages using created_utc from user_data")


        # Check if the user exists in user_data
        if username in user_data:
            # Initialize the criteria list
            criteria_met = []
            logger.info("Checked if the user exists in user_data")

            # Check for characteristics of potential bot accounts
            if total_karma <= KARMA_THRESHOLD:
                criteria_met.append("Low Karma")
            if account_age <= ACCOUNT_AGE_THRESHOLD:
                criteria_met.append("Young Account Age")
                logger.info("Checked for characteristics of potential bot accounts")

            # Check if the user has already been added to the DataFrame
            if username not in added_users:
                # Add the user to the DataFrame
                bot_df = bot_df.append({
                    "User": username,
                    "Average Sentiment Score": average_sentiment_score,
                    "Criteria": ", ".join(criteria_met) if criteria_met else "None",
                    "Link Karma": link_karma,
                    "Comment Karma": comment_karma,
                    "Total Karma": total_karma,
                    "Account Age": account_age,
                    "Awardee Karma": awardee_karma,
                    "User is Contributor": user_is_contributor,
                    "Has Verified Email": has_verified_email,
                    "Accepts Followers": accept_followers,
                },ignore_index=True)
                logger.info(f"Checked if the user has already been added to the DataFrame")

                # Add the user to the set of added users
                added_users.add(username)
                logger.info(f"Added the user to the set of added users")

except Exception as e:
    logger.exception(f"An error occurred while processing submission '{submission_id}' for user '{username}': {e}")


# Specify the Excel file path
EXCEL_FILE_PATH = 'potential_bot_accounts_with_sentiment.xlsx'

# Export the DataFrame to an Excel file
bot_df.to_excel(EXCEL_FILE_PATH, index=False)
logger.debug(f"{bot_df}Excel file created successfully.")

# Load the Excel file using openpyxl
workbook = load_workbook(EXCEL_FILE_PATH)
logger.info(f"Load the Excel file using openpyxl{workbook}")


# Select the active sheet
sheet = workbook.active

# Iterate through the columns and set the column width to the maximum length of data in each column
for i, column_cells in enumerate(sheet.iter_cols(), 1):
    MAX_LENGTH = max(len(str(cell.value)) for cell in column_cells)
    ADJUSTED_WIDTH = MAX_LENGTH + 2  # Add some padding
    sheet.column_dimensions[get_column_letter(i)].width = ADJUSTED_WIDTH

# Save the modified Excel file
workbook.save(EXCEL_FILE_PATH)
logger.debug(f"Saved the modified Excel file:{workbook}")
logger.debug("Potential Bot Accounts:")
logger.debug(bot_df)
logger.info(f"Analysis completed. Identified {bot_df.shape[0]} potential bot accounts and saved to '{EXCEL_FILE_PATH}'.")

