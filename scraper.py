'''
Scrapes top 100 users per karma in the last 3 months
'''
import json
import time
import datetime as dt
import praw
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter  # Import get_column_letter from openpyxl.utils

def load_config():
    """ Load configuration file """
    try:
        with open("config.json", "r", encoding="utf-8") as f:
            config = json.load(f)
    except FileNotFoundError as exc:
        raise FileNotFoundError("The config file 'config.json' was not found. \
            Please create it or check the path.") from exc

    missing = set(["client_id", "client_secret", "subreddit", \
        "username", "password"]) - set(config.keys())
    if len(missing) > 0:
        raise KeyError(f"Missing keys in config.json: {str(missing)}")
    if "user_agent" not in config:
        config["user_agent"] = "Reddit Scraper (by u/Allan_QuartermainSr)"
    return config


def login(config):
    """ Login to Reddit """
    reddit = praw.Reddit(
        user_agent='Reddit Scraper (by u/Allan_QuartermainSr)',
        client_id=config['client_id'],
        client_secret=config['client_secret'],
        password=config['password'],
        username=config['username']
    )
    return reddit

config = load_config()  # Load the configuration
reddit = login(config)  # Login using the config data
subreddit_name = config["subreddit"]
subreddit = reddit.subreddit(subreddit_name)

# Define a dictionary to store user data
user_data = {}

# Define a list to store all submissions without a timeframe
all_submissions = []

# Get the list of moderators
moderators = [mod.name for mod in subreddit.moderator()]

# Check the post sorting method from the config
post_sort = config.get("post_sort", "top")  # Default to "top" if not specified

# List of valid post sorting methods
valid_sort_methods = ["top", "hot", "new", "rising", "controversial"]

# Check if the specified sort method is valid
if post_sort not in valid_sort_methods:
    raise ValueError(f"Invalid post_sort value in config.json. Available \
        options: {', '.join(valid_sort_methods)}")

# Print the note from the config, if available
notes = config.get("notes", "")
if notes:
    print("Note:", notes)
    # Iterate through the posts based on the selected sort method
    if post_sort == "top":
        posts = subreddit.top(limit=10000)
    elif post_sort == "hot":
        posts = subreddit.hot(limit=10000)
    elif post_sort == "new":
        posts = subreddit.new(limit=10000)
    elif post_sort == "rising":
        posts = subreddit.rising(limit=10000)
    elif post_sort == "controversial":
        posts = subreddit.controversial(limit=10000)

# Print a message to indicate scraping has started
print(f'Starting scraping with sort order: {post_sort}')

# Initialize user data
user_data = {}

# Iterate through the posts based on the selected sort method
for submission in posts:
    author = submission.author.name if submission.author else 'Deleted'
    if author in moderators or author == 'Deleted':
        continue
    all_submissions.append(submission)  # Append all submissions

    # Define a list to store submissions within the last 3 months
    submissions_last_3_months = []

    # Calculate the Unix timestamp for three months ago
    three_months_ago = int(time.time()) - 90 * 24 * 60 * 60
# Filter submissions to keep only the ones within the last 3 months
for submission in all_submissions:
    post_date = dt.datetime.fromtimestamp(submission.created_utc)

    # Check if the post is within the last 3 months
    if post_date.timestamp() >= three_months_ago:
        submissions_last_3_months.append(submission)  # Append submissions within the last 3 months

        # Initialize user data if not already present
        author = submission.author.name if submission.author else 'Deleted'
        if author not in user_data:
            user_data[author] = {'User': author, 'Posts': 0, 'Comments': 0, 'Karma': 0}

        # Update user data
        user_data[author]['Posts'] += 1
        user_data[author]['Karma'] += submission.score

        # Retrieve comments for the post
        submission.comments.replace_more(limit=None)
        for comment in submission.comments.list():
            comment_author = comment.author.name if comment.author else 'Deleted'

            # Skip if the comment author is a moderator
            if comment_author in moderators:
                continue

            # Initialize user data if not already present
            if comment_author not in user_data:
                user_data[comment_author] = {'User': comment_author, \
                    'Posts': 0, 'Comments': 0, 'Karma': 0}

            user_data[comment_author]['Comments'] += 1
            user_data[comment_author]['Karma'] += comment.score
# Print a message to indicate scraping is complete
print('Scraping complete.')

# Create a DataFrame from the user data
df = pd.DataFrame(user_data.values())

# Sort the DataFrame by total karma received and limit to top 200
df = df.sort_values(by='Karma', ascending=False).head(200)

# Specify the Excel file path
EXCEL_FILE_PATH = 'user_data.xlsx'

# Export the DataFrame to an Excel file
df.to_excel(EXCEL_FILE_PATH, index=False)

# Load the Excel file using openpyxl
workbook = load_workbook(EXCEL_FILE_PATH)

# Select the active sheet
sheet = workbook.active

# Iterate through the columns and set the column width to the maximum length of data in each column
for i, column in enumerate(sheet.columns):
    MAX_LENGTH = 0
    column = [cell.value for cell in column]
    for value in column:
        if value is not None and len(str(value)) > MAX_LENGTH:
            MAX_LENGTH = len(str(value))
    ADJUSTED_WIDTH = MAX_LENGTH + 2  # Add some padding
    sheet.column_dimensions[get_column_letter(i+1)].width = ADJUSTED_WIDTH

# Save the modified Excel file
workbook.save(EXCEL_FILE_PATH)

print(f'Data written to {EXCEL_FILE_PATH}')
